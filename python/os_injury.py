#!/usr/bin/python3

from openpyxl import Workbook, load_workbook
from tqdm import tqdm
import json
import math

from splitname import splitname
from guessgender import guessgender
#from parsedict import parsedict, prettyEnts, prettyEntsRels
from makedigest import makedigest
from parseyear import parseyear

#dataset_names = { 'ghs':"Guy's Hospital", 'sth':"St Thomas Hospital", 'mxh':'Middlesex Hospital', 'rlh':'Royal London Hospital' }
#dataset_years = { 'ghs':'1810-1839', 'sth':'1773-1809', 'mxh':'1760-1778', 'rlh':'1760-1805' }

dwb = Workbook()
dws = dwb.active
dws.append([

    'person_id',
#    'dp_url',
#    'given',
#    'surname', 
    'gender',
#    'born',
    'agegroup',
#    'description_id',
#    'description_dataset',
#    'description_dataset_name',
#    'description_dataset_years',
#    'description_year',
#    'injury',
    'injury_digest', 
#    'full_description',

#    'archive',
#    'archive_code',
#    'date_of_admission',
#    'name',
#    'reason',
#    'notes',
#    'translation',
#    'discharge_date',
#    'result',
#    'occupation',
#    'abode'

    ])

wb = Workbook()
ws = wb.active
ws.append([

    'person_id',
#    'dp_url',
#    'given',
#    'surname', 
    'gender',
#    'born',
    'agegroup',
#    'description_id',
#    'description_dataset',
#    'description_dataset_name',
#    'description_dataset_years',
#    'description_year',
    'injury', 
    'body_location', 
#    'cause',
#    'full_description', 

#    'place_of_birth',
#    'occupation',
#    'occupation_top',
#    'hisco',
#    'religion',
#    'religion_cat',
#    'married',
#        'has_tattoo',
#    'trials',
#    'earliest_trial_year',
#    'latest_trial_year',
#    'earliest_trial_offence_category',
#    'earliest_trial_sentence_category',
#    'latest_trial_offence_category',
#    'latest_trial_sentencecategory',
#    'earliest_trial_so_type',
#    'latest_trial_so_type',
#    'considered_for_pardon',
#    'penalservitude',
#    'granted_prison_license',
#        'insane',
#        'in_hulks',
#        'transported',
#        'ticket_of_leave',
#        'ship',
#    'executed'

    ])



individuals = {}

def parsesheet(sheet, id_prefix):

    coloffset = 0
    if id_prefix != 'pr': coloffset = 3

    #for i, row in enumerate(tqdm(sheet.iter_rows(min_row=2), total=sheet.max_row, miniters=10, ascii=True, ncols=60)):
    for i, row in enumerate(sheet.iter_rows(min_row=3)):

        #print(i, len(row), row[genderidx].value, row[givenidx].value)
        if len(row) == 0: continue # Required because sometimes openpyxl wants to iterate beyond the end of the data

        #archive = None
        #if archiveidx != -1 and len(row) > archiveidx: archive = row[archiveidx].value

        # We use the archive column to verify that there is actually useful content in the row
        #if archive == None or len(archive) == 0: continue

        individual = id_prefix + '-' + row[0].value + '-' + str(math.floor(row[1].value))

        digest = []

        if individual in individuals: digest = individuals[individual]['digest']

        gender = row[2].value
        agegroup = row[3].value

        inj_element = row[4].value
        side = row[5].value

        rib_location = None
        inj_location = None
        stage_of_healing = None

        if coloffset == 3:
           rib_location = row[6].value
           inj_location = row[7].value
           stage_of_healing = row[8].value

        inj_type = row[6 + coloffset].value
        inj_cat = row[7 + coloffset].value

        inj_count = None
        notes = None
        if len(row) > (9 + coloffset): notes = row[9 + coloffset].value
        if len(row) > (8 + coloffset): inj_count = row[8 + coloffset].value

        #if inj_location:
        #    body = inj_location
        #else:
        #    body = inj_cat

        inj = None
        body = None

        if inj_type.lower() in inj_cat.lower():
            inj = inj_type
            body = inj_cat.lower().replace(inj_type.lower(), '').strip()
        elif 'injury' in inj_cat.lower():
            inj = inj_type
            body = inj_cat.lower().replace('injury', '').strip()
        elif 'dislocation' in inj_cat.lower():
            inj = 'dislocation. ' + inj_type
            body = inj_cat.lower().replace('dislocation', '').strip()
        elif 'fracture' in inj_cat.lower():
            inj = 'fracture. ' + inj_type
            body = inj_cat.lower().replace('fracture', '').strip()
        else:
            print()
            print('**' + inj_type + '**')
            print('**' + inj_cat + '**')
            print()

        if side == 'L':
            body = 'left ' + body
        elif side == 'R':
            body = 'right ' + body
        elif side == 'L&R' or side == 'R and L' or side == 'R&L':
            body = 'left and right ' + body
        elif side == 'U' or side == 'U*' or side == None:
            pass
        else:
            print()
            print('**' + side + '**')
            print()

        if inj_element:
            if inj_element.lower() not in body.lower():
                body = body + ' | ' + inj_element

        if inj_location:
            if inj_location.lower() not in body.lower():
                body = body + ' | ' + inj_location

        if rib_location:
            #print()
            #print('**' + rib_location + '**')
            #print()
            body = rib_location + ' ' + body;
            #print('**' + body + '**')

        #if inj_count:
        #    inj = inj + '. ' + str(math.floor(inj_count))

        if notes:
            inj = inj + ' | ' + notes
            #print()
            #print(inj)

        if stage_of_healing:
            inj = inj + ' | ' + stage_of_healing  

        # we are currently incorporating every field except inj_count. A preliminary look suggests the information is usually repeated elsewhere. Not 100% sure it always is, though.

        #injury = [inj, body, 'INJURYBODY', inj_element, side, rib_location, inj_location, stage_of_healing, inj_type, inj_cat, inj_count, notes]
        injury = [inj, body, 'INJURYBODY']

        ws.append([

            individual, # person_id
            gender, # gender
            agegroup, # agegroup
            inj,  # injury
            body, # body_location

        ])

        digest.append(injury)

        individuals[individual] = {

                'gender': gender,
                'agegroup': agegroup,
                'digest' : digest

        }
        
        #injury = None
        #digest = None
        #if description != None:
        #    description = description.strip() + '.'
        #    ents, rels = parsedict(description)
        #    injury, digest = makedigest(description, ents, rels)

prwb = load_workbook(filename = "data/osteology/Payne Road & Bow Baptist pathology_Revised.xlsx", read_only=True)
prs1 = prwb.worksheets[0]

rlwb = load_workbook(filename = "data/osteology/Royal London Hospital pathology_revised.xlsx", read_only=True)
rls1 = rlwb.worksheets[0]

sbwb = load_workbook(filename = "data/osteology/St. Bride's pathology_revised.xlsx", read_only=True)
sbs1 = sbwb.worksheets[0]


#def parsesheet(sheet, id_prefix):

parsesheet(prs1, 'pr')
parsesheet(rls1, 'rl')
parsesheet(sbs1, 'sb')

for key, individual in individuals.items():

        dws.append([

            key, # lifeid
#            None, # dpurl
#            given, # given
#            surname, # surname
            individual['gender'], # gender
#            born, # born
            individual['agegroup'], # age
#            row_id, # descid
#            dataset_id, # dataset id
#            dataset_names[dataset_id], # dataset names
#            dataset_years[dataset_id], # dataset years
#            description_year, # description year
#            injury, # has injury yes or no
            json.dumps(individual['digest']), # injury digest
#            description, # full description
#
#            archive, # archive
#            row[archivecodeidx].value, # archive_code,
#            row[dateidx].value, # date_of_admission,
#            row[nameidx].value, # name,
#            reason, # reason,
#            notes, # notes,
#            row[descidx].value, # translation,
#            discharge, # discharge_date
#            result, # result
#            occupation, # occupation
#            abode # abode

        ])

dwb.save("data/outputs/os_person.xlsx")
wb.save("data/outputs/os_injury.xlsx")
