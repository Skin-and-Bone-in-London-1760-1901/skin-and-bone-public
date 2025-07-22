#!/usr/bin/python3

import os
import json
from openpyxl import Workbook, load_workbook
from cleandesc import cleandesc
from parsedict import parsedict, prettyEntsRels
from manual import manual
#from getspacytokens import getspacytokens
#from tokenizeents import tokenizeents
#from makejson import makejson
#from makeprodigy import makeprodigy
#from random import randint
from tqdm import tqdm
from makedigest import makedigest
from coldefs import*
from rhclondon import rhclondon

DO_JSON = False
DO_PRODIGY_JSON = False
DO_INJURY_XLSX = True
DO_DESC_XLSX = False
DO_PERSON_XLSX = True
#DO_PEOPLE_LINKING = False
INCLUDE_EMPTY_DESCRIPTIONS = True
TEST_MAX = None
TEST_DESC = None
VERBOSE = False

dataset_names = { 'rhc':'Metropolitan Police Register of Habitual Criminals', 'pld':'UK Licences for Parole of Convicts', 'tlm':'UK Licences for Parole of Male Convicts', 'mpr':'Millbank Prison Register', 'hcr':'England and Wales Criminal Registers' }
dataset_years = { 'rhc':'1881-1925', 'pld':'1853-1925', 'tlm':'1853-1925', 'mpr':'1816-1826', 'hcr':'1791-1802' }

#people = {}

#if (DO_INJURY_XLSX or DO_DESC_XLSX) and DO_PEOPLE_LINKING:
##if False:
#    # load dp_person xlsx file so we can look up person details
#    print('Loading data/outputs/dp_person.xlsx ..')
#    pwb = load_workbook(filename = 'data/outputs/dp_person.xlsx', read_only=True)
#    pws = pwb['Sheet']
#
#
#    print('Processing data/outputs/dp_person.xlsx ..')
##   for i, line in enumerate(tqdm(f, total=total_lines, miniters=500, ascii=True, ncols=60)):
#    for row in tqdm(pws.iter_rows(), total=pws.max_row, miniters=500, ascii=True, ncols=60):
#        people[row[0].value] = row
#
#    pwb.close()

todo = []
#jsonpath = os.path.expanduser("~/digipan/data/larchsbdesc")
jsonpath = os.path.expanduser("~/digipan/data/larchsb")
jsonfiles = os.listdir(jsonpath)
for file in jsonfiles:
  fullpath = jsonpath + "/" + file
  if os.path.isfile(fullpath):
    todo.append(fullpath)

#print(todo)

if DO_INJURY_XLSX:
    wb = Workbook()
    ws = wb.active
    ws.append([

        'person_id',
        'dp_url',
        'given',
        'surname', 
        'gender',
        'born',
        'age',
        'description_id',
        'description_dataset',
        'description_dataset_name',
        'description_dataset_years',
        'description_year',
        'injury', 
        'body_location', 
        'cause',
        'full_description', 

        'place_of_birth',
        'occupation',
        'occupation_top',
        'hisco',
        'religion',
        'religion_cat',
        'married',
#        'has_tattoo',
        'trials',
        'earliest_trial_year',
        'latest_trial_year',
        'earliest_trial_offence_category',
        'earliest_trial_sentence_category',
        'latest_trial_offence_category',
        'latest_trial_sentencecategory',
        'earliest_trial_so_type',
        'latest_trial_so_type',
        'considered_for_pardon',
        'penalservitude',
        'granted_prison_license',
#        'insane',
#        'in_hulks',
#        'transported',
#        'ticket_of_leave',
#        'ship',
        'executed'

        ])

if DO_DESC_XLSX:
    dwb = Workbook()
    dws = dwb.active
    dws.append([

        'person_id',
        'dp_url',
        'given',
        'surname', 
        'gender',
        'born',
        'age',
        'description_id',
        'description_dataset',
        'description_dataset_name',
        'description_dataset_years',
        'description_year',
        'injury',
        'injury_digest', 
        'full_description',

        'place_of_birth',
        'occupation',
        'occupation_top',
        'hisco',
        'religion',
        'religion_cat',
        'married',
#        'has_tattoo',
        'trials',
        'earliest_trial_year',
        'latest_trial_year',
        'earliest_trial_offence_category',
        'earliest_trial_sentence_category',
        'latest_trial_offence_category',
        'latest_trial_sentencecategory',
        'earliest_trial_so_type',
        'latest_trial_so_type',
        'considered_for_pardon',
        'penalservitude',
        'granted_prison_license',
#        'insane',
#        'in_hulks',
#        'transported',
#        'ticket_of_leave',
#        'ship',
        'executed'
 
        ])

if DO_PERSON_XLSX:
    pwb = Workbook()
    pws = pwb.active
    pws.append([

        'person_id',
        'dp_url',
        'given',
        'surname',
        'gender',
        'born',

        'description_count',
        'description_ages',
        'description_ids',
        'description_datasets',
        'description_dataset_names',
        'description_dataset_years',
        'description_years',
        'injury',
        'injury_digests', 
        'full_descriptions',

        'place_of_birth',
        'occupation',
        'occupation_top',
        'hisco',
        'religion',
        'religion_cat',
        'married',
    #    'has_tattoo',
        'trials',
        'earliest_trial_year',
        'latest_trial_year',
        'earliest_trial_offence_category',
        'earliest_trial_sentence_category',
        'latest_trial_offence_category',
        'latest_trial_sentencecategory',
        'earliest_trial_so_type',
        'latest_trial_so_type',
        'considered_for_pardon',
        'penalservitude',
        'granted_prison_license',
    #    'insane',
    #    'in_hulks',
    #    'transported',
    #    'ticket_of_leave',
    #    'ship',
        'executed'
 
        ])


if DO_PRODIGY_JSON:
    prodigyfile = open("./data/outputs/skinandbone_prodigy.jsonl", "w")

if DO_JSON:
    jsonfile = open("./data/outputs/skinandbone.jsonl", "w")

for file in todo:
    
  print('Processing ' + file + ' ..')
  f = open(file, 'r', encoding='utf-8')

  for total_lines, _ in enumerate(f):
      pass
  total_lines = total_lines + 1

  f.seek(0)

  for i, line in enumerate(tqdm(f, total=total_lines, miniters=500, ascii=True, ncols=60)):

    if TEST_MAX and i >= TEST_MAX: break # for testing

    jj = json.loads(line)

    #print(json.dumps(jj, indent=4))
    #exit()

    includeperson = False;

    lifeid = jj['life_id']

    given = None
    surname = None
    gender = None
    born = None
    dpurl = None
    place_of_birth = None
    occupation = None
    occupation_top = None
    hisco = None
    religion = None
    religion_cat = None
    married = None
    trials = None
    earliest_trial_year = None
    latest_trial_year = None
    earliest_trial_offence_category = None
    earliest_trial_sentence_category = None
    latest_trial_offence_category = None
    latest_trial_sentencecategory = None
    earliest_trial_so_type = None
    latest_trial_so_type = None
    considered_for_pardon = None
    penalservitude = None
    granted_prison_license = None
    executed = None

    #j['life_id'],
    dpurl = 'https://www.digitalpanopticon.org/life?id=' + jj['life_id']

    given = jj.get('given', '')
    surname = jj.get('surname', '')
    gender = jj.get('gender')
    born = jj.get('born')
    place_of_birth = splaceofbirth(jj)

    occupation = soccupation(jj)
    occupation_top = soccupationtop(jj)
    hisco = jj.get('hisco')

    religion = sreligion(jj)
    religion_cat = sreligioncat(jj)

    married = jj['global']['sm'].get('married')

#        j['global']['sm'].get('has_tattoo')

    trials = itrials(jj)
    earliest_trial_year = iearliest_trial_year(jj)
    latest_trial_year = ilatest_trial_year(jj)
    earliest_trial_offence_category = searliest_trial_offence_category(jj)
    earliest_trial_sentence_category = searliest_trial_sentence_category(jj)
    latest_trial_offence_category = slatest_trial_offence_category(jj)
    latest_trial_sentencecategory = slatest_trial_sentencecategory(jj)
    earliest_trial_so_type = jj['global']['sm'].get('earliest_trial_so_type')
    latest_trial_so_type = jj['global']['sm'].get('latest_trial_so_type')

    considered_for_pardon = jj['global']['sm'].get('considered_for_pardon')

    penalservitude = jj['global']['sm'].get('penalservitude')
    granted_prison_license = jj['global']['sm'].get('granted_prison_license')

#        j['global']['sm'].get('insane'),
#        j['global']['sm'].get('in_hulks'),
#        j['global']['sm'].get('transported'),
#        j['global']['sm'].get('ticket_of_leave'),
#        sship(j),

    executed = jj['global']['sm'].get('executed')

    person_description_ages = []
    person_description_ids = []
    person_description_datasets = []
    person_description_dataset_names = []
    person_description_dataset_years = []
    person_description_years = []
    person_injury = False
    person_injury_digests = []
    person_full_descriptions = []

#    if (DO_INJURY_XLSX or DO_DESC_XLSX) and DO_PEOPLE_LINKING:
#
#        person = people.get(lifeid)
#        dpurl = person[1].value
#        given = person[2].value
#        surname = person[3].value
#        gender = person[4].value
#        born = person[5].value
#        place_of_birth = person[6].value
#        occupation = person[7].value
#        occupation_top = person[8].value
#        hisco = person[9].value
#        religion = person[10].value
#        religion_cat = person[11].value
#        married = person[12].value
##        has_tattoo = person[12].value
#        trials = person[13].value
#        earliest_trial_year = person[14].value
#        latest_trial_year = person[15].value
#        earliest_trial_offence_category = person[16].value
#        earliest_trial_sentence_category = person[17].value
#        latest_trial_offence_category = person[18].value
#        latest_trial_sentencecategory = person[19].value
#        earliest_trial_so_type = person[20].value
#        latest_trial_so_type = person[21].value
#        considered_for_pardon = person[22].value
#        penalservitude = person[23].value
#        granted_prison_license = person[24].value
#        insane = person[25].value
#        in_hulks = person[26].value
#        transported = person[27].value
#        ticket_of_leave = person[28].value
#        ship = person[29].value
#        executed = person[25].value

    records = jj['records']

    for record in records:

        j = record

        dataset = j['sm']['dataset']
        descid = dataset + j['sm']['id']

        descs = [] # There will not always be a description field. Very occassionally there is more than one.
        desc_year = None # There will not always be a description year
        age = None # There will not always be an age
        include = False

        if (dataset == 'rhc'):

            if rhclondon(jj, j):

                include = True

                desc = j['sm'].get('marks')
                if desc != None: descs.append(desc)

                date = j['dm'].get('date')
                
                if (date):
                    desc_year = date['ld']['year']

                    if (desc_year > 1901):
                        include = False

                age = None
                if (born and desc_year):
                    age = desc_year - born

        if (dataset == 'pld_fen'):

            if j['sm']['citable_reference'].startswith('PCOM 4/'):

                include = True

                desc = j['sm'].get('dist_marks') 
                desc2 = j['sm'].get('recep_dist_marks')
                desc3 = j['sm'].get('rel_dist_marks')

                if desc != None: descs.append(desc)
                if desc2 != None and desc2 != desc: descs.append(desc2)
                if desc3 != None and desc3 != desc and desc3 != desc2: descs.append(desc3)

                desc_year = j['sm'].get('comm_year')

                if desc_year != None:
                    desc_year = int(desc_year)
                    if (desc_year > 1901):
                        include = False

                age = j['sm'].get('age')


        if (dataset == 'tlm_pen'):

            include = True

            desc = j['sm'].get('distinctive_marks')
            desc2 = j['sm'].get('rec_distinctive_marks')
            desc3 = j['sm'].get('rel_distinctive_marks')

            if desc != None: descs.append(desc)
            if desc2 != None and desc2 != desc: descs.append(desc2)
            if desc3 != None and desc3 != desc and desc3 != desc2: descs.append(desc3)
            
            desc_year = j['sm'].get('comm_year')

            #date = j['dm'].get('comm_date')
                
            #if (date):
            #    desc_year = date['ld']['year']

            if (desc_year == None): desc_year = j['sm'].get('conv_year')

                #print('desc_year was none')
                #print(j)
                #exit()

                #date = j['dm'].get('conv_date')
                    
                #if (date):
                #    desc_year = date['ld']['year']

            if desc_year != None:
                desc_year = int(desc_year)
                if (desc_year > 1901):
                    print('TLM_PEN TOO LATE')
                    include = False

            age = j['sm'].get('age_yrs')

        #if (dataset == 'tlm_med'):
        #
        #    desc = j['sm'].get('wound_and_other_injuries')
        #    desc_year = j['sm'].get('date_of_comittal')
        #    age = j['sm'].get('age')

        if (dataset == 'mpr'):

            include = True

            desc = j['sm'].get('Appearance_marks')
            if desc != None: descs.append(desc)

            date = j['dm'].get('date')
            
            if (date):
                desc_year = date['ld']['year']

            age = j['sm'].get('age')

        if (dataset == 'hcr'):

            include = True

            desc = j['sm'].get('description')
            if desc != None: descs.append(desc)

            date = j['dm'].get('date')
            
            if (date):
                desc_year = date['ld']['year']

            age = j['sm'].get('age')
          

        if include and len(descs) == 0 and INCLUDE_EMPTY_DESCRIPTIONS: descs.append('');

        if include:

            includeperson = True
            #print(descs)

            for desc in descs:

                if TEST_DESC:
                    desc = TEST_DESC

                #print('-----------------------------------------------------------------------------------------------------------')
                #print(desc)

                description = cleandesc(desc)
                ents, rels = manual(descid)
                if ents == False:
                    ents, rels = parsedict(description)
                #else:
                #    print('MANUAL:')
                #    print(ents, rels)
                #    ents, rels = parsedict(description)
                #    print('PARSEDICT:')
                #    print(ents, rels)
                #    exit()

                if VERBOSE and rels and len(rels) > 0: prettyEntsRels(ents, rels, description)

                if rels != None:

                    if DO_JSON:

                        jsn = makejson(description, descid, ents, rels)
                        jsonfile.write(json.dumps(jsn) + '\n')

                    if DO_PRODIGY_JSON:

                        spacytokens = getspacytokens(description)
                        ents = tokenizeents(ents, spacytokens)

                        prodigy = makeprodigy(description, spacytokens, ents, rels)

                        # The rel_component tutorial uses the meta:source property to divide the data sets.
                        # This is a hack obviously and needs replacing with a proper solution.
                        meta = {}
                        random = randint(1, 100)        
                        if random < 33:
                            meta['source'] = 'BioNLP 2011 Genia Shared Task, PMC-2806624-00-TIAB.txt'
                        elif random < 66:
                            meta['source'] = 'BioNLP 2011 Genia Shared Task, PMC-1134653-00-TIAB.txt'
                        else:
                            meta['source'] = 'BioNLP 2011 Genia Shared Task, PMC-1134658-00-TIAB.txt'
                        prodigy['meta'] = meta
                        # End of hack

                        prodigyfile.write(json.dumps(prodigy) + '\n')

                    if DO_INJURY_XLSX:

                        toadds = []

                        # This has to be altered:
                        #     - iterate through the ents
                        #     - if ent is an INJURY:
                        #           - iterate through the rels
                        #               - does rel involve this injury and is an INJURYCAUSE?
                        #                   - make a note of the injury cause
                        #           - iterate through the rels
                        #               - does rel involve this injury and is an INJURYBODY?
                        #                   - Add row with this INJURY and BODYLOCATION
                        #           - Was the ent not involved in any INJURYBODY rels?
                        #               - Add row with this INJURY and no BODYLOCATION

                        for ent in ents:

                            if ent[2] == 'INJURY':
                                
                                in_injurybody = False
                                causeent = None

                                for head, child, label in rels:

                                    headent = ents[head]
                                    childent = ents[child]

                                    if label == 'INJURYCAUSE' and headent == ent:

                                        causeent = childent

                                for head, child, label in rels:

                                    headent = ents[head]
                                    childent = ents[child]

                                    if label == 'INJURYBODY' and headent == ent:

                                        toadds.append((ent, childent, causeent))
                                        in_injurybody = True

                                if in_injurybody == False:

                                    toadds.append((ent, None, causeent))

                        for injuryent, bodyent, causeent in toadds:

                            bodytext = None
                            if bodyent: bodytext = description[bodyent[0]:bodyent[1]]

                            causetext = None
                            if causeent: causetext = description[causeent[0]:causeent[1]]

                            ws.append([

                                lifeid,
                                dpurl,
                                given,
                                surname,
                                gender,
                                born,
                                age,
                                descid,
                                descid[:3],
                                dataset_names[descid[:3]],
                                dataset_years[descid[:3]],
                                desc_year,
                                description[injuryent[0]:injuryent[1]],
                                bodytext,
                                causetext,
                                desc,

                                place_of_birth,
                                occupation,
                                occupation_top,
                                hisco,
                                religion,
                                religion_cat,
                                married,
        #                        has_tattoo,
                                trials,
                                earliest_trial_year,
                                latest_trial_year,
                                earliest_trial_offence_category,
                                earliest_trial_sentence_category,
                                latest_trial_offence_category,
                                latest_trial_sentencecategory,
                                earliest_trial_so_type,
                                latest_trial_so_type,
                                considered_for_pardon,
                                penalservitude,
                                granted_prison_license,
        #                        insane,
        #                        in_hulks,
        #                        transported,
        #                        ticket_of_leave,
        #                        ship,
                                executed

                            ])

                injury, digest = makedigest(description, ents, rels)

                person_description_ages.append(age)
                person_description_ids.append(descid)
                person_description_datasets.append(descid[:3])
                person_description_dataset_names.append(dataset_names[descid[:3]])
                person_description_dataset_years.append(dataset_years[descid[:3]])
                person_description_years.append(desc_year)
                if injury: person_injury = True
                person_injury_digests.append(digest)
                person_full_descriptions.append(desc)

                if DO_DESC_XLSX:

                    #injury = 'no'
                    #digest = []

                    #if rels and len(rels) > 0:
                    #    injury = 'yes'
                    #    digest = makedigest(description, ents, rels)

                    dws.append([

                        lifeid,
                        dpurl,
                        given,
                        surname,
                        gender,
                        born,
                        age,
                        descid,
                        descid[:3],
                        dataset_names[descid[:3]],
                        dataset_years[descid[:3]],
                        desc_year,
                        injury,
                        json.dumps(digest),
                        desc,

                        place_of_birth,
                        occupation,
                        occupation_top,
                        hisco,
                        religion,
                        religion_cat,
                        married,
        #                has_tattoo,
                        trials,
                        earliest_trial_year,
                        latest_trial_year,
                        earliest_trial_offence_category,
                        earliest_trial_sentence_category,
                        latest_trial_offence_category,
                        latest_trial_sentencecategory,
                        earliest_trial_so_type,
                        latest_trial_so_type,
                        considered_for_pardon,
                        penalservitude,
                        granted_prison_license,
        #                insane,
        #                in_hulks,
        #                transported,
        #                ticket_of_leave,
        #                ship,
                        executed

                        ])

    if DO_PERSON_XLSX and includeperson:

        pws.append([

            lifeid,
            dpurl,
            given,
            surname,
            gender,
            born,

            len(person_full_descriptions),
            str(person_description_ages),
            str(person_description_ids),
            str(person_description_datasets),
            str(person_description_dataset_names),
            str(person_description_dataset_years),
            str(person_description_years),
            person_injury,
            str(person_injury_digests), 
            str(person_full_descriptions),

            place_of_birth,
            occupation,
            occupation_top,
            hisco,
            religion,
            religion_cat,
            married,
    #                has_tattoo,
            trials,
            earliest_trial_year,
            latest_trial_year,
            earliest_trial_offence_category,
            earliest_trial_sentence_category,
            latest_trial_offence_category,
            latest_trial_sentencecategory,
            earliest_trial_so_type,
            latest_trial_so_type,
            considered_for_pardon,
            penalservitude,
            granted_prison_license,
    #                insane,
    #                in_hulks,
    #                transported,
    #                ticket_of_leave,
    #                ship,
            executed

        ])

    i += 1

if DO_JSON:
    jsonfile.close()

if DO_PRODIGY_JSON:
    prodigyfile.close()

if DO_INJURY_XLSX:
    print('Saving data/outputs/dp_injury.xlsx ..')
    wb.save("data/outputs/dp_injury.xlsx")

if DO_DESC_XLSX:
    print('Saving data/outputs/dp_desc.xlsx ..')
    dwb.save("data/outputs/dp_desc.xlsx")

if DO_PERSON_XLSX:
    print('Saving data/outputs/dp_person.xlsx ..')
    pwb.save("data/outputs/dp_person.xlsx")
