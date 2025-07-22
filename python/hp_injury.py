#!/usr/bin/python3

from openpyxl import Workbook, load_workbook
from tqdm import tqdm
import json

from splitname import splitname
from guessgender import guessgender
from parsedict import parsedict, prettyEnts, prettyEntsRels
from manual import manual
from makedigest import makedigest
from parseyear import parseyear

TEST_DESC = None
VERBOSE = False

dataset_names = { 'ghs':"Guy's Hospital", 'sth':"St Thomas Hospital", 'mxh':'Middlesex Hospital', 'rlh':'Royal London Hospital' }
dataset_years = { 'ghs':'1810-1839', 'sth':'1773-1809', 'mxh':'1760-1778', 'rlh':'1760-1805' }

dwb = Workbook()
dws = dwb.active
dws.append([

    'person_id',
    'dp_url',
    'given',
    'surname', 
    'gender',
    'born',
    'age',
    'description_id',
    'description_dataset',
    'description_dataset_name',
    'description_dataset_years',
    'description_year',
    'injury',
    'injury_digest', 
    'full_description',

    'archive',
    'archive_code',
    'date_of_admission',
    'name',
    'reason',
    'notes',
    'translation',
    'discharge_date',
    'result',
    'occupation',
    'abode'

    ])

wb = Workbook()
ws = wb.active
ws.append([

    'person_id',
    'dp_url',
    'given',
    'surname', 
    'gender',
    'born',
    'age',
    'description_id',
    'description_dataset',
    'description_dataset_name',
    'description_dataset_years',
    'description_year',

    'injury', 
    'body_location', 
    'cause',

    'full_description',

    'archive',
    'archive_code',
    'date_of_admission',
    'name',
    'reason',
    'notes',
    'translation',
    'discharge_date',
    'result',
    'occupation',
    'abode'

    ])


def parsesheet(sheet, dataset_id, dataset_sheet, descidx, nameidx, dateidx, archiveidx, archivecodeidx, reasonidx, notesidx, ageidx, dischargeidx, resultidx, occupationidx, abodeidx, givenidx, surnameidx, yearidx, genderidx):

    for i, row in enumerate(tqdm(sheet.iter_rows(min_row=2), total=sheet.max_row, miniters=10, ascii=True, ncols=60)):
    #for i, row in enumerate(sheet.iter_rows(min_row=2)):

        #print(i, len(row), row[genderidx].value, row[givenidx].value)
        if len(row) == 0: continue # Required because sometimes openpyxl wants to iterate beyond the end of the data

        archive = None
        if archiveidx != -1 and len(row) > archiveidx: archive = row[archiveidx].value

        # We use the archive column to verify that there is actually useful content in the row
        if archive == None or len(archive) == 0: continue

        row_id = dataset_id + '-' + str(dataset_sheet) + '-' + str(i + 1)

        reason = None
        if reasonidx != -1 and len(row) > reasonidx: reason = row[reasonidx].value

        notes = None
        if notesidx != -1 and len(row) > notesidx: notes = row[notesidx].value

        discharge = None
        if dischargeidx != -1 and len(row) > dischargeidx: discharge = row[dischargeidx].value

        result = None
        if resultidx != -1 and len(row) > resultidx: result = row[resultidx].value

        occupation = None
        if occupationidx != -1 and len(row) > occupationidx: occupation = row[occupationidx].value

        abode = None
        if abodeidx != -1 and len(row) > abodeidx: abode = row[abodeidx].value

        given = None
        if givenidx != -1 and len(row) > givenidx: given = row[givenidx].value

        surname = None
        if surnameidx != -1 and len(row) > surnameidx: surname = row[surnameidx].value

        if nameidx != -1 and len(row) > nameidx and given == None:
            given, surname = splitname(row[nameidx].value)

        gender = None
        if genderidx != -1 and len(row) > genderidx: gender = row[genderidx].value

        if gender == None and given != None:
            gender = guessgender(given)

        admission_date = None
        if dateidx != -1 and len(row) > dateidx: admission_date = row[dateidx].value;

        description_year = None
        if yearidx != -1 and len(row) > yearidx: description_year = row[yearidx].value

        if description_year == None and admission_date != None:
            description_year = int(parseyear(admission_date))

        age = None
        if ageidx != -1 and len(row) > ageidx: age = row[ageidx].value

        if age != None:
            if isinstance(age, float):
                age = int(age)
            elif age.isdigit():
                age = int(age)
            else:
                age = None

        born = None
        if age and description_year: born = description_year - age
       
        description = None
        if descidx != -1 and len(row) > descidx: description = row[descidx].value

        injury = None
        digest = None
        if description != None:

            if TEST_DESC:
                description = TEST_DESC

            description = description.lower().strip() + '.'
            descid = 'hp_' + row_id
            ents, rels = manual(descid)
            if ents == False:
                ents, rels = parsedict(description)

            if VERBOSE:
                print()
                print()
                prettyEnts(ents, description)
                prettyEntsRels(ents, rels, description)
                print()
                if TEST_DESC: exit()

            toadds = []

            # This has to be altered:
            #     - iterate through the ents
            #     - if ent is an INJURY:
            #           - iterate through the rels
            #               - does rel involve this injury and is an INJURYCAUSE?
            #                   - make a note of the injury cause
            #           - iterate through the rels
            #               - does rel involve this injury and is an INJURYBODY?
            #                   - Add row with this INJURY and BODYLOCATION
            #           - Was the ent not involved in any INJURYBODY rels?
            #               - Add row with this INJURY and no BODYLOCATION

            for ent in ents:

                if ent[2] == 'INJURY':
                    
                    in_injurybody = False
                    causeent = None

                    for head, child, label in rels:

                        headent = ents[head]
                        childent = ents[child]

                        if label == 'INJURYCAUSE' and headent == ent:

                            causeent = childent

                    for head, child, label in rels:

                        headent = ents[head]
                        childent = ents[child]

                        if label == 'INJURYBODY' and headent == ent:

                            toadds.append((ent, childent, causeent))
                            in_injurybody = True

                    if in_injurybody == False:

                        toadds.append((ent, None, causeent))

            for injuryent, bodyent, causeent in toadds:

                bodytext = None
                if bodyent: bodytext = description[bodyent[0]:bodyent[1]]

                causetext = None
                if causeent: causetext = description[causeent[0]:causeent[1]]

                ws.append([

                    row_id, # lifeid
                    None, # dpurl
                    given, # given
                    surname, # surname
                    gender, # gender
                    born, # born
                    age, # age
                    row_id, # descid
                    dataset_id, # dataset id
                    dataset_names[dataset_id], # dataset names
                    dataset_years[dataset_id], # dataset years
                    description_year, # description year

                    description[injuryent[0]:injuryent[1]], # injury 
                    bodytext, # body_location 
                    causetext, # cause

                    description, # full description

                    archive, # archive
                    row[archivecodeidx].value, # archive_code,
                    row[dateidx].value, # date_of_admission,
                    row[nameidx].value, # name,
                    reason, # reason,
                    notes, # notes,
                    row[descidx].value, # translation,
                    discharge, # discharge_date
                    result, # result
                    occupation, # occupation
                    abode # abode

                ])


#                ws.append([
#
#                    lifeid,
#                    dpurl,
#                    given,
#                    surname,
#                    gender,
#                    born,
#                    age,
#                    descid,
#                    descid[:3],
#                    dataset_names[descid[:3]],
#                    dataset_years[descid[:3]],
#                    desc_year,
#                    description[injuryent[0]:injuryent[1]],
#                    bodytext,
#                    causetext,
#                    desc,
#
#                    place_of_birth,
#                    occupation,
#                    occupation_top,
#                    hisco,
#                    religion,
#                    religion_cat,
#                    married,
##                        has_tattoo,
#                    trials,
#                    earliest_trial_year,
#                    latest_trial_year,
#                    earliest_trial_offence_category,
#                    earliest_trial_sentence_category,
#                    latest_trial_offence_category,
#                    latest_trial_sentencecategory,
#                    earliest_trial_so_type,
#                    latest_trial_so_type,
#                    considered_for_pardon,
#                    penalservitude,
#                    granted_prison_license,
##                        insane,
##                        in_hulks,
##                        transported,
##                        ticket_of_leave,
##                        ship,
#                    executed
#
#                ])






















            injury, digest = makedigest(description, ents, rels)

        #if 'by a dog' in description or i == 0:
        #    print()
        #    prettyEntsRels(ents, rels, description)
        #    print(injury, digest)
        #    prettyEnts(ents, description)

        #if rels and len(rels) > 0: prettyEntsRels(ents, rels, description)
        #if description != 'burns.'\
        #and description != 'bruises.'\
        #and description != 'private injury.'\
        #and description != 'internal injury.'\
        #and description != 'scalded.'\
        #and description != 'sprain.'\
        #and description != 'burn.'\
        #and description != 'bruised.'\
        #and description != 'contusion.'\
        #and description != 'contusions.'\
        #and description != 'scalds.'\
        #and description != 'fractured.'\
        #and description != 'violent sprain.'\
        #and description != 'dislocated.'\
        #and description != 'fracture.'\
        #and description != 'lacerated.'\
        #and description != 'scald.':
        #    if rels is None or len(rels) == 0:
        #        print()
        #        print('***' + description + '***')
        #        print(injury, digest)
        #        #prettyEntsRels(ents, rels, description)
        #        prettyEnts(ents, description)

        dws.append([

            row_id, # lifeid
            None, # dpurl
            given, # given
            surname, # surname
            gender, # gender
            born, # born
            age, # age
            row_id, # descid
            dataset_id, # dataset id
            dataset_names[dataset_id], # dataset names
            dataset_years[dataset_id], # dataset years
            description_year, # description year
            injury, # has injury yes or no
            json.dumps(digest), # injury digest
            description, # full description

            archive, # archive
            row[archivecodeidx].value, # archive_code,
            row[dateidx].value, # date_of_admission,
            row[nameidx].value, # name,
            reason, # reason,
            notes, # notes,
            row[descidx].value, # translation,
            discharge, # discharge_date
            result, # result
            occupation, # occupation
            abode # abode

        ])






ghwb = load_workbook(filename = "data/hospital/Guy's Hospital.xlsx", read_only=True)
ghs1 = ghwb['1810-1819 Injuries']
ghs2 = ghwb['1820-1829 Injuries']
ghs3 = ghwb['1830-1839 Injuries']

stwb = load_workbook(filename = "data/hospital/St. Thomas' Hospital.xlsx", read_only=True)
sts1 = stwb['1773-6 Injuries']
sts2 = stwb['1781-9 Injuries']
sts3 = stwb['1790-9 Injuries']
sts4 = stwb['1800-9 Injuries']

mxwb = load_workbook(filename = "data/hospital/Middlesex Hospital.xlsx", read_only=True)
mxs1 = mxwb['Injuries']

rlwb = load_workbook(filename = "data/hospital/Royal London Hospital.xlsx", read_only=True)
rls1 = rlwb['1760 Injuries']
rls2 = rlwb['1791-2 Injuries']
rls3 = rlwb['1805 Injuries']

#def parsesheet(sheet, dataset_id, dataset_sheet, descidx, nameidx, dateidx, archiveidx, archivecodeidx, reasonidx, notesidx, ageidx, dischargeidx, resultidx, occupationidx, abodeidx, givenidx, surnameidx, yearidx, genderidx):

parsesheet(ghs1, 'ghs', 1, 6, 3, 2, 0, 1, 4, 5, -1, -1, -1, -1, -1, -1, -1, -1, -1)
parsesheet(ghs2, 'ghs', 2, 5, 3, 2, 0, 1, -1, 4, -1, -1, -1, -1, -1, -1, -1, -1, -1)
parsesheet(ghs3, 'ghs', 3, 6, 3, 2, 0, 1, -1, 5, 4, -1, -1, -1, -1, -1, -1, -1, -1)

parsesheet(sts1, 'sth', 1, 5, 3, 2, 0, 1, 4, 6, -1, -1, -1, -1, -1, -1, -1, -1, -1)
parsesheet(sts2, 'sth', 2, 5, 3, 2, 0, 1, 4, 6, -1, -1, -1, -1, -1, -1, -1, -1, -1)
parsesheet(sts3, 'sth', 3, 5, 3, 2, 0, 1, 4, 6, -1, -1, -1, -1, -1, -1, -1, -1, -1)
parsesheet(sts4, 'sth', 4, 5, 3, 2, 0, 1, 4, 6, -1, -1, -1, -1, -1, -1, -1, -1, -1)

parsesheet(mxs1, 'mxh', 1, 5, 3, 2, 0, 1, -1, -1, -1, 6, 7, -1, -1, -1, -1, -1, -1)

parsesheet(rls1, 'rlh', 1, 13, -1, 4, 0, 1, -1, -1, 11, 14, 16, 10, 9, 6, 7, 5, 8)
parsesheet(rls2, 'rlh', 2, 13, -1, 4, 0, 1, -1, -1, 11, 14, 16, 10, 9, 6, 7, 5, 8)
parsesheet(rls3, 'rlh', 3, 11, -1, 4, 0, 1, -1, -1, -1, 12, 14, 10, 9, 6, 7, 5, 8)

dwb.save("data/outputs/hp_person.xlsx")
wb.save("data/outputs/hp_injury.xlsx")
